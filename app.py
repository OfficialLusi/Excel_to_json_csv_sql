# app_gui.py

import tkinter as tk  # GUI toolkit
from tkinter import ttk, filedialog, messagebox  # widgets and dialogs
from pathlib import Path  # FS paths
from typing import Dict, List  # typing

HeaderPattern = ["COLUMN 1", "COLUMN 2", "...", "COLUMN N"]  # header pattern 

class App(tk.Tk):
    def __init__(self):
        super().__init__()  # init tk 
        self.title("Excel → Tables/DDL")  # window title 
        self.geometry("820x640")  # window size 

        self.ExcelPathVar = tk.StringVar(value="")  # chosen excel path 
        self.OutDirVar = tk.StringVar(value=str(Path("out").resolve()))  # output directory 
        self.JsonVar = tk.BooleanVar(value=True)  # export JSON flag 
        self.CsvVar = tk.BooleanVar(value=True)  # export CSV flag 
        self.SqlVar = tk.BooleanVar(value=True)  # export SQL flag 
        self.StopOnEmptyRowVar = tk.BooleanVar(value=True)  # stop on empty row flag 
        self.HeaderPatternVar = tk.StringVar(value=", ".join(HeaderPattern))  # header pattern

        self.Sheets: List[str] = []  # available sheet names 
        self.SheetListVar = tk.Variable(value=self.Sheets)  # listbox var 
        self.SchemaBySheet: Dict[str, tk.StringVar] = {}  # per-sheet schema vars 

        self._BuildUi()  # build UI 

    def _BuildUi(self):
        pad = {"padx": 8, "pady": 6}  # padding 

        # File selection row #
        fileFrame = ttk.LabelFrame(self, text="Source Excel")  # frame 
        fileFrame.pack(fill="x", **pad)  # layout 
        ttk.Entry(fileFrame, textvariable=self.ExcelPathVar).pack(side="left", expand=True, fill="x", padx=6, pady=6)  # path entry 
        ttk.Button(fileFrame, text="Browse...", command=self._ChooseExcel).pack(side="left", padx=6, pady=6)  # browse button 
        ttk.Button(fileFrame, text="Load Sheets", command=self._LoadSheets).pack(side="left", padx=6, pady=6)  # load sheets 

        # Output options #
        outFrame = ttk.LabelFrame(self, text="Output")  # frame 
        outFrame.pack(fill="x", **pad)  # layout 
        ttk.Checkbutton(outFrame, text="JSON", variable=self.JsonVar).pack(side="left", padx=6, pady=6)  # json flag 
        ttk.Checkbutton(outFrame, text="CSV", variable=self.CsvVar).pack(side="left", padx=6, pady=6)  # csv flag 
        ttk.Checkbutton(outFrame, text="SQL (DDL)", variable=self.SqlVar).pack(side="left", padx=6, pady=6)  # sql flag 
        ttk.Checkbutton(outFrame, text="Stop on empty row", variable=self.StopOnEmptyRowVar).pack(side="left", padx=6, pady=6)  # stop flag 

        # Output directory #
        outDirFrame = ttk.Frame(outFrame)  # nested frame 
        outDirFrame.pack(fill="x", padx=6, pady=6)  # layout 
        ttk.Label(outDirFrame, text="Output dir:").pack(side="left")  # label 
        ttk.Entry(outDirFrame, textvariable=self.OutDirVar).pack(side="left", expand=True, fill="x", padx=6)  # path entry 
        ttk.Button(outDirFrame, text="Choose...", command=self._ChooseOutDir).pack(side="left")  # choose dir 

        # Header pattern # 
        patternFrame = ttk.LabelFrame(self, text="Header Pattern (defines table start)")  # frame
        patternFrame.pack(fill="x", **pad)  # layout
        ttk.Label(patternFrame, text="Set Frame (format: col_name_a, col_name_b, ....): ").pack(side="left", padx=6, pady=6)  # label
        ttk.Entry(patternFrame, textvariable=self.HeaderPatternVar).pack(side="left", expand=True, fill="x", padx=6, pady=6)  # entry

        # Sheets + schemas #
        sheetsFrame = ttk.LabelFrame(self, text="Sheets & Schemas")  # frame 
        sheetsFrame.pack(fill="both", expand=True, **pad)  # layout 

        left = ttk.Frame(sheetsFrame)  # left pane 
        left.pack(side="left", fill="both", expand=True)  # layout 
        ttk.Label(left, text="Available sheets (select which to parse):").pack(anchor="w", padx=6, pady=4)  # label 
        self.SheetList = tk.Listbox(left, listvariable=self.SheetListVar, selectmode="extended", height=12)  # listbox 
        self.SheetList.pack(fill="both", expand=True, padx=6, pady=4)  # layout #

        right = ttk.Frame(sheetsFrame)  # right pane 
        right.pack(side="left", fill="both", expand=True)  # layout 
        ttk.Label(right, text="Schemas by sheet (default used if left blank):").pack(anchor="w", padx=6, pady=4)  # label 
        self.SchemaCanvas = tk.Canvas(right, borderwidth=0)  # canvas for scroll 
        self.SchemaScroll = ttk.Scrollbar(right, orient="vertical", command=self.SchemaCanvas.yview)  # scrollbar 
        self.SchemaFrame = ttk.Frame(self.SchemaCanvas)  # inner frame 
        self.SchemaFrame.bind("<Configure>", lambda e: self.SchemaCanvas.configure(scrollregion=self.SchemaCanvas.bbox("all")))  # update scroll 
        self.SchemaCanvas.create_window((0, 0), window=self.SchemaFrame, anchor="nw")  # place frame 
        self.SchemaCanvas.configure(yscrollcommand=self.SchemaScroll.set)  # connect scroll 
        self.SchemaCanvas.pack(side="left", fill="both", expand=True, padx=6, pady=4)  # layout 
        self.SchemaScroll.pack(side="left", fill="y")  # layout #

        # Action buttons #
        actionFrame = ttk.Frame(self)  # bottom actions 
        actionFrame.pack(fill="x", **pad)  # layout 
        ttk.Button(actionFrame, text="Run", command=self._Run).pack(side="right", padx=6)  # run 
        ttk.Button(actionFrame, text="Quit", command=self.destroy).pack(side="right", padx=6)  # quit 

    def _ChooseExcel(self):
        p = filedialog.askopenfilename(title="Choose Excel file", filetypes=[("Excel", "*.xlsx *.xlsm *.xltx *.xltm")])  # file dialog 
        if p:
            self.ExcelPathVar.set(p)  # set chosen path 

    def _ChooseOutDir(self):
        p = filedialog.askdirectory(title="Choose output directory")  # directory dialog 
        if p:
            self.OutDirVar.set(p)  # set chosen dir 

    def _LoadSheets(self):
        try:
            excelPath = Path(self.ExcelPathVar.get())  # read path 
            if not excelPath.exists():
                messagebox.showerror("Error", "Excel file not found.")  # error 
                return
            from openpyxl import load_workbook  # lazy import 
            wb = load_workbook(excelPath, read_only=True)  # open workbook 
            self.Sheets = list(wb.sheetnames)  # fetch sheet names 
            self.SheetListVar.set(self.Sheets)  # update listbox 
            self._BuildSchemaEntries(self.Sheets)  # build schema entries 
            messagebox.showinfo("Sheets loaded", f"Found {len(self.Sheets)} sheets.")  # info 
        except Exception as e:
            messagebox.showerror("Error", str(e))  # show error 

    def _BuildSchemaEntries(self, sheets: List[str]):
        # clear previous 
        for child in list(self.SchemaFrame.children.values()):
            child.destroy()  # remove old widgets 
        self.SchemaBySheet.clear()  # reset map 
        # create row with label + entry for each sheet 
        for i, name in enumerate(sheets):
            lbl = ttk.Label(self.SchemaFrame, text=name)  # sheet label 
            lbl.grid(row=i, column=0, sticky="w", padx=4, pady=2)  # place 
            var = tk.StringVar(value="")  # schema var 
            ent = ttk.Entry(self.SchemaFrame, textvariable=var, width=24)  # schema entry 
            ent.grid(row=i, column=1, sticky="w", padx=4, pady=2)  # place 
            self.SchemaBySheet[name] = var  # store var 

    def _SelectedSheets(self) -> List[str]:
        idx = list(self.SheetList.curselection())  # selected indices 
        if not idx:  # nothing selected 
            return self.Sheets  # default: all sheets 
        return [self.Sheets[i] for i in idx]  # map to names 

    def _EffectiveSchemaBySheet(self, selectedSheets: List[str]) -> Dict[str, str]:
        # take provided per-sheet schema or use default from the first non-empty value or fallback "MYSCHEMA" 
        candidates = [v.get().strip() for v in self.SchemaBySheet.values() if v.get().strip()]  # non-empty schemas 
        defaultSchema = candidates[0] if candidates else "MYSCHEMA"  # default schema 
        mapping: Dict[str, str] = {}  # result 
        for s in selectedSheets:
            v = self.SchemaBySheet.get(s).get().strip() if self.SchemaBySheet.get(s) else ""  # sheet schema 
            mapping[s] = v or defaultSchema  # per-sheet or default 
        return mapping  # map 

    # def _HandleTablesHeader(self, header: List[str]) -> List[str]: # get a new defined header
        

    def _Run(self):
        try:
            excelPath = Path(self.ExcelPathVar.get())  # excel path 
            outDir = Path(self.OutDirVar.get())  # output dir 
            if not excelPath.exists():
                messagebox.showerror("Error", "Excel file not found.")  # error 
                return
            outDir.mkdir(parents=True, exist_ok=True)  # ensure dir 
            selectedSheets = self._SelectedSheets()  # sheets to parse 
            schemaBySheet = self._EffectiveSchemaBySheet(selectedSheets)  # schema map 

            if self.HeaderPatternVar.get() is not None or self.HeaderPatternVar.get().strip() != "":
                HeaderPattern = [s.strip() for s in self.HeaderPatternVar.get().split(",") if s.strip()]  # parse pattern

            from extractor_module import ExtractAllTables, WriteTablesJson, WriteTablesCsv 
            
            tables = ExtractAllTables(excelPath, HeaderPattern, sheetsToCheck=set(selectedSheets), stopOnEmptyRow=self.StopOnEmptyRowVar.get())  # parse tables 
            if not tables:
                messagebox.showwarning("No tables", "No tables found with the given header pattern.")  # warn 
                return
            # Exports #
            if self.JsonVar.get():
                from extractor_module import WriteTablesJson
                WriteTablesJson(tables, outDir / "tables.json")  # write JSON 
            if self.CsvVar.get():
                csvDir = outDir / "csv"  # csv folder 
                from extractor_module import WriteTablesCsv
                WriteTablesCsv(tables, csvDir)  # write CSVs 
            if self.SqlVar.get():
                sqlDir = outDir / "sql"  # sql folder 
                from ddl_module import WriteAllDdls  # import here to avoid cycles 
                WriteAllDdls(tables, schemaBySheet, sqlDir)  # write DDLs 
            messagebox.showinfo("Done", f"Completed.\nOutput: {outDir.resolve()}")  # success 
        except Exception as e:
            messagebox.showerror("Error", str(e))  # error dialog 

if __name__ == "__main__":
    App().mainloop()  # run app 

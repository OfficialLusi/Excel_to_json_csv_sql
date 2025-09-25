# extractor_module.py

import json  # JSON read/write
from pathlib import Path  # FS paths
from typing import List, Tuple, Dict, Any, Optional  # typing
from openpyxl import load_workbook  # Excel reader
from openpyxl.worksheet.worksheet import Worksheet  # typing hint

def NormalizeString(value: Any) -> str:
    value = "" if value is None else str(value).strip()  # normalize None to empty and trim spaces 
    value = " ".join(value.split())  # collapse multiple spaces 
    return value.casefold()  # case-insensitive compare 

def FindHeaderPositions(sheet: Worksheet, pattern: List[str]) -> List[Tuple[int, int]]:
    normalizedPattern = [NormalizeString(x) for x in pattern]  # normalize pattern tokens 
    columnsCount = len(normalizedPattern)  # number of header cells 
    positions: List[Tuple[int, int]] = []  # collection of header hits 
    maxRow = sheet.max_row  # openpyxl estimated last row 
    maxCol = sheet.max_column  # openpyxl estimated last column 
    for row in range(1, maxRow + 1):  # scan rows 
        for col in range(1, maxCol - columnsCount + 2):  # scan viable columns 
            window = [NormalizeString(sheet.cell(row, col + k).value) for k in range(columnsCount)]  # read window 
            if window == normalizedPattern:  # match found 
                positions.append((row, col))  # store top-left header cell 
    return positions  # all positions 

def ResolveMergedCellValue(sheet: Worksheet, row: int, col: int):
    for merged in sheet.merged_cells.ranges:  # iterate merged ranges 
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:  # inside 
            return sheet.cell(merged.min_row, merged.min_col).value  # return anchor value 
    return sheet.cell(row, col).value  # not merged 

def GetTableName(sheet: Worksheet, headerRow: int, headerCol: int, columnsCount: int) -> str:
    titleRow = headerRow - 1  # table title expected one row above 
    if titleRow >= 1:  # row valid 
        for col in range(headerCol, headerCol + columnsCount):  # scan above header width 
            value = ResolveMergedCellValue(sheet, titleRow, col)  # merged-aware read 
            if value and str(value).strip():  # first non-empty wins 
                return str(value).strip()  # use as table name 
    return sheet.title  # fallback to sheet name 

def RowIsEmpty(sheet: Worksheet, row: int, startCol: int, endCol: int) -> bool:
    for col in range(startCol, endCol + 1):  # scan columns in row 
        value = sheet.cell(row, col).value  # read cell 
        if value is not None and str(value).strip() != "":  # found non-empty 
            return False  # not empty 
    return True  # empty row 

def ReadTable(sheet: Worksheet, headerStart: Tuple[int, int], pattern: List[str], stopOnEmptyRow: bool = True) -> Dict[str, Any]:
    row0, col0 = headerStart  # header top-left 
    columnsCount = len(pattern)  # header size 
    tableName = GetTableName(sheet, row0, col0, columnsCount)  # infer table name 
    header = [str(sheet.cell(row0, col0 + k).value).strip() for k in range(columnsCount)]  # header labels 
    rows: List[Dict[str, Any]] = []  # collected rows 
    row = row0 + 1  # first data row 
    maxRow = sheet.max_row  # bottom bound 
    while row <= maxRow:  # scan downward 
        if RowIsEmpty(sheet, row, col0, col0 + columnsCount - 1):  # empty row in window 
            if stopOnEmptyRow:  # stop policy 
                break  # end table 
            row += 1  # skip empty row 
            continue  # next row 
        rowValues = [sheet.cell(row, col0 + k).value for k in range(columnsCount)]  # read row window 
        if [NormalizeString(v) for v in rowValues] == [NormalizeString(v) for v in pattern]:  # next header 
            break  # new table starts here 
        rows.append({header[i]: rowValues[i] for i in range(columnsCount)})  # header→value mapping 
        row += 1  # next row 
    return {  # table object 
        "table_name": tableName,  # table name 
        "header": header,  # header labels 
        "rows": rows,  # data rows 
        "sheet": sheet.title,  # source sheet 
        "start_cell": (row0, col0),  # header position (row, col) 1-based 
    }

def ExtractAllTables(excelPath: Path, pattern: List[str], sheetsToCheck: Optional[set] = None, stopOnEmptyRow: bool = True) -> List[Dict[str, Any]]:
    workBook = load_workbook(excelPath, data_only=True)  # open workbook with computed values 
    targetSheets = [n for n in workBook.sheetnames if (sheetsToCheck is None or n in sheetsToCheck)]  # filter sheets 
    tables: List[Dict[str, Any]] = []  # accumulator 
    for name in targetSheets:  # each sheet 
        sheet = workBook[name]  # get worksheet 
        headerPositions = FindHeaderPositions(sheet, pattern)  # locate header(s) 
        for start in headerPositions:  # each header 
            tables.append(ReadTable(sheet, start, pattern, stopOnEmptyRow=stopOnEmptyRow))  # parse table 
    return tables  # all tables 

def WriteTablesJson(tables: List[Dict[str, Any]], outputPath: Path):
    outputPath.write_text(json.dumps(tables, ensure_ascii=False, indent=2), encoding="utf-8")  # write JSON 

def WriteTablesCsv(tables: List[Dict[str, Any]], outDir: Path):
    outDir.mkdir(parents=True, exist_ok=True)  # ensure folder 
    import csv  # CSV writer 
    for t in tables:  # one CSV per table 
        tableName = t["table_name"].replace("/", "_").replace("\\", "_").replace(" ", "_")  # filename-safe name 
        csvPath = outDir / f"{tableName}.csv"  # target path 
        header = t["header"]  # header order 
        with open(csvPath, "w", newline="", encoding="utf-8") as f:  # open file 
            writer = csv.DictWriter(f, fieldnames=header)  # writer with header 
            writer.writeheader()  # header row 
            for row in t["rows"]:  # each row 
                writer.writerow(row)  # write row 